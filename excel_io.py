from __future__ import annotations
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Optional

import openpyxl


# ===== Excel Layout (wie VBA) =====
SHEET_EINGABE = "Anpassung"
COL_EMP = 13       # M
COL_PROJ = 14      # N
COL_ABS = 16       # P
FIRST_ROW_LIST = 4

HEADER_ROW = 3
SUBHEADER_ROW = 4
FIRST_EMP_COL = 6  # F
DATE_FIRST_ROW = 5
DATE_COL = 3       # C

H1 = 3.5
H2 = 7.0


MONTH_DE = {
    1: "Januar",
    2: "Februar",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Dezember",
}


def month_sheet_name(dt: date) -> str:
    return MONTH_DE[dt.month]


def _as_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _normalize_key(s: str) -> str:
    return (s or "").strip().lower()


@dataclass
class EmployeeBlock:
    start_col: int
    width: int

    @property
    def abs_col(self) -> int:
        return self.start_col + self.width - 1


class ExcelIO:
    """
    Öffnet .xlsm mit keep_vba=True und schreibt Werte ähnlich zu deinem VBA.
    """

    def __init__(self, file_path: str, retries: int = 3, retry_wait_sec: float = 1.2):
        self.file_path = file_path
        self.retries = retries
        self.retry_wait_sec = retry_wait_sec

    # -------------------------
    # Public: Listen aus "Anpassung"
    # -------------------------
    def load_lists(self) -> tuple[list[str], list[str], list[str]]:
        wb = self._open_workbook()
        try:
            if SHEET_EINGABE not in wb.sheetnames:
                raise RuntimeError(f"Blatt '{SHEET_EINGABE}' nicht gefunden.")
            ws = wb[SHEET_EINGABE]
            emps = self._unique_from_col(ws, COL_EMP, FIRST_ROW_LIST)
            projs = self._unique_from_col(ws, COL_PROJ, FIRST_ROW_LIST)
            abss = self._unique_from_col(ws, COL_ABS, FIRST_ROW_LIST)
            return emps, projs, abss
        finally:
            wb.close()

    # -------------------------
    # Public: Kalender-Markierung (gefüllte Tage)
    # -------------------------
    def get_filled_days_for_employee(self, emp: str, month: date) -> set[date]:
        """
        Liefert alle Tage (Mo-Fr + Wochenende egal), an denen im Monatsblatt
        für den Mitarbeiter-Block irgendeine Zelle (Projekt oder Abwesenheit) gefüllt ist.
        """
        emp = (emp or "").strip()
        if not emp:
            return set()

        wb = self._open_workbook()
        try:
            ws = self._get_month_sheet(wb, month)
            if ws is None:
                return set()

            block = self._find_employee_block(ws, emp)
            if block is None:
                return set()

            filled: set[date] = set()
            max_row = ws.max_row

            for r in range(DATE_FIRST_ROW, max_row + 1):
                dv = _as_date(ws.cell(r, DATE_COL).value)
                if not dv:
                    continue
                # Nur aktueller Monat
                if dv.year != month.year or dv.month != month.month:
                    continue

                # Check alle Zellen im Mitarbeiterblock (inkl. Abs-Spalte)
                for c in range(block.start_col, block.start_col + block.width):
                    v = ws.cell(r, c).value
                    if v not in (None, "", 0):
                        filled.add(dv)
                        break

            return filled
        finally:
            wb.close()

    # -------------------------
    # Public: Schreiben (mehrere Tage)
    # -------------------------
    def write_range(
        self,
        emp: str,
        mode: str,        # "PROJ" oder "ABS"
        proj: str,
        hrs: float,
        abs_type: str,
        d_from: date,
        d_to: date,
    ) -> tuple[int, int]:
        """
        Schreibt Mo-Fr im Zeitraum, analog VBA.
        returns (ok_count, fail_count)
        """
        emp = (emp or "").strip()
        proj = (proj or "").strip()
        abs_type = (abs_type or "").strip()

        if d_to < d_from:
            d_from, d_to = d_to, d_from

        wb = self._open_workbook()
        ok = 0
        fail = 0
        try:
            cur = d_from
            while cur <= d_to:
                if cur.weekday() <= 4:  # Mo=0 .. Fr=4
                    if self._write_one_day(wb, cur, emp, mode, proj, hrs, abs_type):
                        ok += 1
                    else:
                        fail += 1
                cur += timedelta(days=1)

            wb.save(self.file_path)
        finally:
            wb.close()

        return ok, fail

    # -------------------------
    # Internal: Workbook öffnen mit Retry
    # -------------------------
    def _open_workbook(self):
        last_err = None
        for _ in range(self.retries):
            try:
                return openpyxl.load_workbook(self.file_path, keep_vba=True)
            except Exception as e:
                last_err = e
                time.sleep(self.retry_wait_sec)
        raise RuntimeError(f"Excel-Datei konnte nicht geöffnet werden (evtl. gesperrt): {last_err}")

    # -------------------------
    # Internal: Monatsblatt holen
    # -------------------------
    def _get_month_sheet(self, wb, month: date):
        nm = month_sheet_name(month)
        if nm in wb.sheetnames:
            return wb[nm]
        if nm == "März" and "Maerz" in wb.sheetnames:
            return wb["Maerz"]
        return None

    # -------------------------
    # Internal: Unique List aus Spalte
    # -------------------------
    def _unique_from_col(self, ws, col: int, first_row: int) -> list[str]:
        out = []
        seen = set()
        max_row = ws.max_row
        for r in range(first_row, max_row + 1):
            v = ws.cell(r, col).value
            s = (str(v).strip() if v is not None else "")
            if s and s not in seen:
                seen.add(s)
                out.append(s)
        return out or []

    # -------------------------
    # Internal: 1 Tag schreiben (wie VBA)
    # -------------------------
    def _write_one_day(
        self,
        wb,
        dt: date,
        emp: str,
        mode: str,
        proj: str,
        hrs: float,
        abs_type: str,
    ) -> bool:
        ws = self._get_month_sheet(wb, dt)
        if ws is None:
            return False

        day_row = self._find_date_row(ws, dt)
        if day_row == 0:
            return False

        block = self._find_employee_block(ws, emp)
        if block is None:
            return False

        abs_col = block.abs_col

        if mode == "ABS":
            ws.cell(day_row, abs_col).value = abs_type
            for c in range(block.start_col, abs_col):
                ws.cell(day_row, c).value = None
            return True

        proj_col = self._find_project_col(ws, block, proj)
        if proj_col == 0:
            return False

        ws.cell(day_row, proj_col).value = float(hrs)
        ws.cell(day_row, abs_col).value = None
        return True

    # -------------------------
    # Internal: Datumszeile finden (Spalte C)
    # -------------------------
    def _find_date_row(self, ws, dt: date) -> int:
        max_row = ws.max_row
        target = dt
        for r in range(DATE_FIRST_ROW, max_row + 1):
            dv = _as_date(ws.cell(r, DATE_COL).value)
            if dv and dv == target:
                return r
        return 0

    # -------------------------
    # Internal: Mitarbeiter-Block finden (Zeile 3, ggf. Merges)
    # -------------------------
    def _find_employee_block(self, ws, emp: str) -> Optional[EmployeeBlock]:
        emp_key = _normalize_key(emp)
        c = FIRST_EMP_COL
        empty_streak = 0
        max_c = ws.max_column

        while c <= max_c:
            name, width, next_c = self._header_cell_value_and_width(ws, HEADER_ROW, c)
            name_key = _normalize_key(name)

            if name_key:
                empty_streak = 0
                if name_key == emp_key:
                    return EmployeeBlock(start_col=c, width=width)
            else:
                empty_streak += 1

            c = next_c
            if empty_streak >= 15:
                break

        return None

    def _header_cell_value_and_width(self, ws, row: int, col: int) -> tuple[str, int, int]:
        cell = ws.cell(row, col)

        merged_range = None
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                merged_range = rng
                break

        if merged_range:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            val = top_left.value
            width = merged_range.max_col - merged_range.min_col + 1
            start_col = merged_range.min_col
            next_col = start_col + width
            return (str(val).strip() if val is not None else ""), width, next_col

        val = cell.value
        return (str(val).strip() if val is not None else ""), 1, col + 1

    # -------------------------
    # Internal: Projektspalte finden (Zeile 4 innerhalb Block, ohne Abs-Spalte)
    # -------------------------
    def _find_project_col(self, ws, block: EmployeeBlock, proj: str) -> int:
        proj_key = _normalize_key(proj)
        for c in range(block.start_col, block.start_col + block.width - 1):
            v = ws.cell(SUBHEADER_ROW, c).value
            if _normalize_key(str(v) if v is not None else "") == proj_key:
                return c
        return 0
